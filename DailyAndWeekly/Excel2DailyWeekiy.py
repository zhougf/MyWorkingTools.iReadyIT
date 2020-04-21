# python
# -*- coding:utf-8 -*-

"""
@FileName: Excel2DailyWeekiy.py
@Version: v1.0
@Author: Micheal Zhou
@CreateTime: 2020-04-07 14:10
@License: GPL
@Contact: zhougf930@163.com
@See:
"""

import datetime, time, io, sys, os, traceback, json, re, string
from math import ceil
from openpyxl import *


# 通过当前时间计算出所在当月周次，进而获得需要的Excel表名
def GetSheetName(whichWeek):
    num = ['一', '二', '三', '四', '五', '六']
    if whichWeek == "this":
        dt = datetime.datetime.now()
    if whichWeek == "next":
        dt = datetime.datetime.now() + datetime.timedelta(days=7)
    first_day = dt.replace(day=1)
    dom = dt.day
    adjusted_dom = dom + first_day.weekday()
    print(int(ceil(adjusted_dom / 7.0)))
    sheetTitle = time.strftime('%Y.%m', time.localtime(time.time())) + "第" + num[
        int(ceil(adjusted_dom / 7.0)) - 1] + "周"
    print(sheetTitle + "\n")
    return sheetTitle


# 根据Excel路径和表名参数，获取该sheet表中的数据，并转化为整理后的二维数组
def Excel22array(excel_path, sheetTitle):
    list22array = []
    # 打开文件
    wb = load_workbook(excel_path)
    # 打印所有表名
    # print(wb.sheetnames)
    # for sheet in wb:
    #    print(sheet.title)
    # print("\n")
    # 读取表数据
    ws = wb[sheetTitle]
    # ws = wb.get_sheet_by_name(nTitle)
    # 获取最大行
    row_max = ws.max_row
    # print(row_max)
    # print("\n")
    # 获取最大列
    # con_max = ws.max_column
    # print(con_max)
    # 把上面写入数组的内容打印在控制台
    for m in range(1, row_max + 1):
        list = []
        for n in range(1, 14):
            list.append(ws.cell(row=m, column=n).value)
        list22array.append(list)
    for m in range(1, row_max):
        if list22array[m][1] == None:
            list22array[m][1] = list22array[m - 1][1]
            # print(m + 1)
            # print(list22array[m][1])
    # print("\n")
    return list22array


# 将计划任务的二维数组转化为每个人的当天/指定日期工作列表和全部当天任务保存到制定位置的文件中
def Array2DailyString(todayStr, list22array, nameList, txt_path):
    dailyString = ""
    allDailyString = ""
    if todayStr == "":
        todayStr = str(time.strftime('%Y-%m-%d', time.localtime(time.time())))
    dateMark = "*" + todayStr + "*"
    if list22array != [] and nameList != []:
        for m in range(4, 11):
            if todayStr in str(list22array[1][m]):
                for k in range(len(nameList)):
                    iNo = 0
                    taskStr = ""
                    for n in range(2, len(list22array)):
                        if nameList[k] in list22array[n][3] and list22array[n][m] != None and str(
                                list22array[n][11]) != "已取消":
                            iNo = iNo + 1
                            taskStr = taskStr + str(iNo) + "*" + str(list22array[n][1]) + "*" + str(
                                list22array[n][2]) + "*" + str(format(list22array[n][m], '.0%')) + "\n"
                    if taskStr != "":
                        dailyString = dateMark + str(nameList[k]) + "\n" + taskStr
                        print(dailyString)
                        allDailyString = allDailyString + dailyString + "\n"
                        txt_path_and_filename = txt_path + todayStr + str(nameList[k]) + ".txt"
                        Write2DailyFile(txt_path_and_filename, dailyString)
    if allDailyString != "":
        txt_path_and_filename = txt_path + todayStr + "AllDaily.txt"
        Write2DailyFile(txt_path_and_filename, allDailyString)


# 将二维数组转化为用于填写企业微信日报的文本
def Array2DailyString4QYWX(todayStr, list22array, projectList, nameList, txt_path):
    allDailyString = ""
    if todayStr == "":
        todayStr = str(time.strftime('%Y-%m-%d', time.localtime(time.time())))
    dateMark = "*" + todayStr + "*"
    if list22array != [] and nameList != []:
        for m in range(4, 11):
            if todayStr in str(list22array[1][m]):
                for k in range(len(nameList)):
                    iNo = 0
                    dailyString = ""
                    for p in range(len(projectList)):
                        taskStr = ""
                        for n in range(2, len(list22array)):
                            if projectList[p] in list22array[n][1] and nameList[k] in list22array[n][3] and \
                                    list22array[n][m] != None and str(list22array[n][11]) != "已取消":
                                iNo = iNo + 1
                                taskStr = taskStr + str(iNo) + "*" + str(list22array[n][1]) + "*" + str(
                                    list22array[n][2]) + "*" + str(format(list22array[n][m], '.0%')) + "\n"
                        if taskStr != "":
                            dailyString = dailyString + dateMark + "\n" + taskStr
                    if dailyString != "":
                        dailyString = nameList[k] + "\n" + dailyString + "\n"
                        allDailyString = allDailyString + dailyString
                        print(dailyString)
                        txt_path_and_filename = txt_path + todayStr + str(nameList[k]) + ".txt"
                        Write2DailyFile(txt_path_and_filename, dailyString)
    if allDailyString != "":
        txt_path_and_filename = txt_path + todayStr + "AllDaily.txt"
        Write2DailyFile(txt_path_and_filename, allDailyString)


# 将二维数组转化为用于同步到有道笔记的日报文本
def Array2DailyString4YDNote(todayStr, list22array, name):
    dailyString = ""
    if todayStr == "":
        todayStr = str(time.strftime('%Y-%m-%d', time.localtime(time.time())))
    dateMark = "*" + todayStr + "*"
    if list22array != [] and name != "":
        for m in range(4, 11):
            if todayStr in str(list22array[1][m]):
                iNo = 0
                taskStr = ""
                for n in range(2, len(list22array)):
                    if list22array[n][m] != None and name in str(list22array[n][3]) and str(
                            list22array[n][11]) != "已取消":
                        iNo = iNo + 1
                        taskStr = taskStr + str(iNo) + "*" + str(list22array[n][1]) + "*" + str(
                            list22array[n][2]) + "*" + str(format(list22array[n][m], '.0%')) + "\n"
                if taskStr != "":
                    dailyString = dateMark + "\n" + taskStr
                    # print(dailyString)
    return dailyString


# 将整理后的日报文字写到特定的文件里
def Write2DailyFile(txt_path_and_filename, dailyString):
    with io.open(txt_path_and_filename, "wb") as txt:
        txt.writelines(dailyString)


# 将二维数组中有用的信息转化为周报内容并保存到文件中
def Array2WeeklyString(thisweekly, tw2array, nextweekly, nw2array, projectList, txt_path):
    weeklyString = ""
    pNo = 0
    if projectList != []:
        tWeekly = ""
        for p in range(len(projectList)):
            iNo = 0
            projtask = ""
            for n in range(2, len(tw2array)):
                if projectList[p] in tw2array[n][1] and str(tw2array[n][11]) != "已取消":
                    iNo = iNo + 1
                    for k in range(4, 11):
                        if tw2array[n][k] != None:
                            taskProgress = str(format(tw2array[n][k], '.0%'))
                    # print(taskProgress)
                    projtask = projtask + "（" + str(iNo) + "）" + str(tw2array[n][2]) + "*" + taskProgress + "\n"
            if projtask != "":
                pNo = pNo + 1
                projtask = str(pNo) + "." + str(projectList[p]) + "\n" + projtask
                tWeekly = tWeekly + projtask
        weeklyString = "【" + thisweekly + "】" + tw2array[2][12] + "\n" + tWeekly + "\n"
        pNo = 0
        nWeekly = ""
        for p in range(len(projectList)):
            iNo = 0
            projtask = ""
            for n in range(2, len(nw2array)):
                if projectList[p] in nw2array[n][1]:
                    iNo = iNo + 1
                    projtask = projtask + "（" + str(iNo) + "）" + str(nw2array[n][2]) + "\n"
            if projtask != "":
                pNo = pNo + 1
                projtask = str(pNo) + "." + str(projectList[p]) + "\n" + projtask
                nWeekly = nWeekly + projtask
        weeklyString = weeklyString + "【" + nextweekly + "】" + nw2array[2][12] + "\n" + nWeekly
    if weeklyString != "":
        print(weeklyString)
        txt_path_and_filename = txt_path + "Weekly.txt"
        Write2DailyFile(txt_path_and_filename, weeklyString)


# 程序执行的入口
if __name__ == "__main__":
    reload(sys)
    sys.setdefaultencoding('utf-8')
    os.environ['NLS_LANG'] = 'Simplified Chinese_CHINA.ZHS16GBK'

    todayStr = ""
    projectList = ['计划与管理', '技能和业务培训', '海尔项目', '三机O2O', '北汽二期', '摩根华鑫', '瑞穗银行', '宝钢气体', '贝克曼MS-Flow']
    nameList = ['白明晨', '曹珊', '丛震', '董杰', '段语', '高思佳', '李茂清', '陶仙', '王虎林', '王羽超', '杨彦刚', '周光甫']
    excel_path = "/Users/michealzhou/MyDriver/工作/06部门管理/02工作计划与汇报/周报/上海项目计划与周报-20200420.xlsx"  # 设置Excel文件路径
    txt_path = "/Users/michealzhou/Desktop/"  # 设置text文件路径

    try:
        thisweekSheetname = GetSheetName("this")
        # nextweekSheetname = GetSheetName("next")
        # thisweekSheetname = "2020.04第三周"
        # nextweekSheetname = "2020.04第四周"
        tw2array = Excel22array(excel_path, thisweekSheetname)
        # nw2array = Excel22array(excel_path, nextweekSheetname)

        Array2DailyString4QYWX(todayStr, tw2array, projectList, nameList, txt_path)
        print(Array2DailyString4YDNote(todayStr, tw2array, "周光甫"))

        # Array2DailyString(todayStr, tw2array, nameList, txt_path)
        # Array2WeeklyString(thisweekSheetname, tw2array, nextweekSheetname, nw2array, projectList, txt_path)

    except Exception as e:
        traceback.print_exc()
